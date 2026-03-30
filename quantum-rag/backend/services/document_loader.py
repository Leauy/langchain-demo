"""Document loader service for various file types."""
import os
import tempfile
from typing import List, Tuple
from pathlib import Path
from langchain_core.documents import Document

import openpyxl
from pypdf import PdfReader


class DocumentLoader:
    """Service for loading documents from various file types."""

    def load_file(self, file_path: str, file_type: str) -> List[Document]:
        """
        Load documents from a file.

        Args:
            file_path: Path to the file
            file_type: Type of file (excel, pdf, txt, markdown)

        Returns:
            List of Document objects
        """
        loaders = {
            "excel": self._load_excel,
            "xlsx": self._load_excel,
            "xls": self._load_excel,
            "pdf": self._load_pdf,
            "txt": self._load_text,
            "markdown": self._load_text,
            "md": self._load_text,
        }

        loader = loaders.get(file_type.lower())
        if not loader:
            raise ValueError(f"Unsupported file type: {file_type}")

        return loader(file_path)

    def _load_excel(self, file_path: str) -> List[Document]:
        """Load documents from Excel file."""
        documents = []
        workbook = openpyxl.load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else "")

            # Process data rows
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and cell.value:
                        row_data[headers[col_idx]] = str(cell.value)

                if row_data:
                    # Determine module/sub_module from common column names
                    module = row_data.get("模块", row_data.get("module", sheet_name))
                    sub_module = row_data.get("子模块", row_data.get("sub_module", ""))

                    # Build content from all columns
                    content_parts = []
                    for key, value in row_data.items():
                        if key not in ["模块", "子模块", "module", "sub_module"]:
                            content_parts.append(f"{key}: {value}")

                    content = "\n".join(content_parts)

                    if content.strip():
                        doc = Document(
                            page_content=content,
                            metadata={
                                "module": module,
                                "sub_module": sub_module,
                                "source": f"{sheet_name}!A{row_idx}",
                                "row": row_idx
                            }
                        )
                        documents.append(doc)

        return documents

    def _load_pdf(self, file_path: str) -> List[Document]:
        """Load documents from PDF file."""
        documents = []
        reader = PdfReader(file_path)

        for page_num, page in enumerate(reader.pages, start=1):
            text = page.extract_text()
            if text.strip():
                doc = Document(
                    page_content=text,
                    metadata={
                        "module": f"Page {page_num}",
                        "sub_module": "",
                        "source": f"page_{page_num}",
                        "page": page_num
                    }
                )
                documents.append(doc)

        return documents

    def _load_text(self, file_path: str) -> List[Document]:
        """Load documents from text/markdown file."""
        documents = []

        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Split by sections (headers in markdown or double newlines)
        sections = content.split("\n## ")

        file_name = Path(file_path).stem

        for i, section in enumerate(sections):
            if not section.strip():
                continue

            lines = section.strip().split("\n")
            title = lines[0].replace("#", "").strip() if lines else f"Section {i+1}"
            section_content = "\n".join(lines[1:]) if len(lines) > 1 else section

            if section_content.strip():
                doc = Document(
                    page_content=section_content,
                    metadata={
                        "module": file_name,
                        "sub_module": title,
                        "source": f"section_{i+1}"
                    }
                )
                documents.append(doc)

        # If no sections found, create single document
        if not documents and content.strip():
            doc = Document(
                page_content=content,
                metadata={
                    "module": file_name,
                    "sub_module": "Full Document",
                    "source": "full"
                }
            )
            documents.append(doc)

        return documents


# Singleton instance
document_loader = DocumentLoader()
